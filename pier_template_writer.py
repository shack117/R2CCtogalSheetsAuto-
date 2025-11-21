# pier_template_writer.py

from pathlib import Path
from typing import Dict, Tuple

from openpyxl import load_workbook
from pier_logic import PierMetrics

# Row label keywords found in column A of the pier section on BID sheet
ROW_LABELS = {
    # Row 12: "Shaft Dia inches"
    "shaft_dia_in": ["SHAFT DIA", "SHAFT", "SHAFT DIAMETER"],

    # Row 13: "Bell Dia inches (if no bell, bell=shaft dia)"
    "bell_dia_in": ["BELL DIA", "BELL DIAMETER", "BELL"],

    # Row 14: "Pier Depth LF"
    "depth_ft": ["PIER DEPTH", "DEPTH"],

    # Row 15: "Pier Qty"
    "count": ["PIER QTY", "QTY", "QTY."],

    # We are NOT writing total_length_lf (row 29 has formulas),
    # but keep keywords here for future logging if needed.
    "total_length_lf": ["TOTAL LENGTH LF PIERS", "TOTAL LENGTH", "TOTAL LF"],
}


def _find_row(sheet, keywords):
    """
    Search column A for the first row whose text contains ANY keyword
    (case-insensitive). Returns the row index or None.
    """
    if not keywords:
        return None

    for r in range(1, sheet.max_row + 1):
        val = sheet.cell(row=r, column=1).value
        if not val:
            continue

        text = str(val).upper()
        for kw in keywords:
            if kw in text:
                return r
    return None


def _make_safe_sheet_title(tier_name: str) -> str:
    """
    Excel sheet names cannot contain: : \ / ? * [ ]
    and must be <= 31 chars.
    """
    forbidden = set(r':\/?*[]')
    cleaned = "".join(ch if ch not in forbidden else "_" for ch in tier_name)
    cleaned = cleaned.strip()
    if not cleaned:
        cleaned = "TIER"
    return cleaned[:31]


def _get_template_sheet(wb):
    """
    Return the base sheet we should clone per tier.
    For your template this is the 'BID' sheet.
    We DO NOT rename or delete BID so that formulas pointing to BID stay valid.
    """
    if "BID" in wb.sheetnames:
        return wb["BID"]
    # fallback: first sheet if BID was renamed in the future
    return wb[wb.sheetnames[0]]


def _create_tier_sheets(wb, tiers):
    """
    Given a workbook and a list of tier names (strings), create one sheet
    per tier by copying the BID sheet as a template.

    IMPORTANT: We DO NOT touch the original BID sheet.
    """
    template_sheet = _get_template_sheet(wb)
    tier_to_sheet = {}

    for tier in tiers:
        safe_title = _make_safe_sheet_title(tier)
        sheet = wb.copy_worksheet(template_sheet)
        sheet.title = safe_title
        tier_to_sheet[tier] = sheet

    return tier_to_sheet


def write_piers_to_template(
    template_path: str,
    output_path: str,
    pier_metrics: Dict[Tuple[str, str], PierMetrics],
):
    """
    Layer 3: write drilled pier metrics into the estimating template.

    - Loads template (.xlsx or .xlsm)
      * Keeps VBA only if both template and output are .xlsm
    - Detects unique tiers in pier_metrics
    - Creates one 'BID-style' worksheet per tier by copying BID
    - For each tier sheet:
        - writes pier condition names into row 5 starting at column C
        - writes shaft, bell, depth, and count into the appropriate rows
    - DOES NOT overwrite rows that hold formulas like 'Total Length LF Piers'
    """
    print("\n=== Layer 3: Writing Pier Data into Template ===")
    print(f"Loading template: {template_path}")

    template_suffix = Path(template_path).suffix.lower()
    output_suffix = Path(output_path).suffix.lower()

    # Decide whether to preserve VBA:
    # Only safe/valid when both template and output are .xlsm
    keep_vba = template_suffix == ".xlsm" and output_suffix == ".xlsm"

    if keep_vba:
        print("Loading workbook with VBA preservation (keep_vba=True).")
        wb = load_workbook(template_path, data_only=False, keep_vba=True)
    else:
        print("Loading workbook without VBA preservation (keep_vba=False).")
        wb = load_workbook(template_path, data_only=False)

    # Collect unique tiers from the metrics (e.g., 'TIER 1', 'TIER 2', 'UNASSIGNED')
    tiers = sorted({tier for (tier, _) in pier_metrics.keys()})
    if not tiers:
        print("No tiers found in pier metrics; nothing to write.")
        wb.save(output_path)
        print(f"Saved (unchanged) template as: {output_path}")
        return

    print(f"Detected tiers: {tiers}")

    # Create one cloned BID-style sheet per tier
    tier_to_sheet = _create_tier_sheets(wb, tiers)

    # Group metrics by tier so we can assign columns left-to-right
    tier_buckets: Dict[str, list[PierMetrics]] = {t: [] for t in tiers}
    for (tier, cls), metrics in pier_metrics.items():
        # Only true drilled piers here; skip Pier Caps ("PC - ...")
        if not cls.upper().startswith("PIER"):
            continue
        if tier in tier_buckets:
            tier_buckets[tier].append(metrics)

    # Write each tier sheet
    for tier, metrics_list in tier_buckets.items():
        sheet = tier_to_sheet[tier]
        print(f"\n--- Writing Tier: {tier} into sheet '{sheet.title}' ---")

        if not metrics_list:
            print("  No drilled piers for this tier.")
            continue

        # Sort conditions by classification so columns are in a stable order
        metrics_list.sort(key=lambda m: m.classification)

        start_col = 3  # column C
        header_row = 5

        for idx, metrics in enumerate(metrics_list):
            col = start_col + idx

            # Header: "PIER - 1", "PIER - 2", etc.
            header_cell = sheet.cell(row=header_row, column=col)
            header_cell.value = metrics.classification

            print(f"  Condition '{metrics.classification}' → column {col} ({header_cell.coordinate})")

            # Map metric names → values
            field_map = {
                "shaft_dia_in": metrics.shaft_dia_in,
                "bell_dia_in": metrics.bell_dia_in,
                "depth_ft": metrics.depth_ft,
                "count": metrics.count,
                # "total_length_lf": metrics.total_length_lf,  # leave formulas intact
            }

            # Write values into the correct rows under this column
            for field, value in field_map.items():
                if value is None:
                    continue

                keywords = ROW_LABELS.get(field, [])
                row = _find_row(sheet, keywords)
                if not row:
                    print(f"    [WARN] Could not find row for {field} (keywords={keywords})")
                    continue

                cell = sheet.cell(row=row, column=col)
                cell.value = value
                print(f"    Wrote {field} = {value} → {cell.coordinate}")

    wb.save(output_path)
    print(f"\nSaved completed estimate: {output_path}\n")
